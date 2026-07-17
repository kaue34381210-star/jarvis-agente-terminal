"""Ferramentas de projeto, documentos, memória e terminal."""
import os
import glob
import json
import re
import shlex
import tempfile
import datetime
import subprocess

from . import caminhos
from . import config
from . import permissao


_IGNORAR = {".git", "node_modules", ".venv", "venv", "__pycache__",
            ".mypy_cache", ".pytest_cache", ".ruff_cache", "dist", "build",
            ".next", ".turbo", "target", ".idea", ".gradle"}

_ASSINATURAS_BINARIAS = (
    (b"\x89PNG", "PNG"),
    (b"%PDF", "PDF"),
    (b"PK\x03\x04", "ZIP/container"),
    (b"PK\x05\x06", "ZIP/container"),
    (b"PK\x07\x08", "ZIP/container"),
    (b"\x7fELF", "ELF"),
)
_BYTES_CONTROLE = frozenset(range(32)) - {9, 10, 13}
_BYTES_CONTROLE = _BYTES_CONTROLE | {127}


def _dentro(base: str, caminho: str) -> str:
    return caminhos.exigir_dentro(base, caminho)


def _resolver_alvo(caminho: str) -> str:
    """Resolve um alvo de escrita com a mesma regra do gate de risco."""
    return caminhos.resolver(config.REPO, caminho)


def _garantir() -> None:
    os.makedirs(config.DADOS, exist_ok=True)


def _arquivo_resumo() -> str:
    return os.path.join(config.DADOS, "memoria_resumo.txt")


def _truncar(texto: str, limite: int, unidade: str = "chars") -> str:
    """Trunca texto e informa exatamente o volume omitido."""
    if len(texto) <= limite:
        return texto
    omitidos = len(texto) - limite
    return (f"{texto[:limite]}\n...[truncado: {omitidos} {unidade} omitidos "
            f"de {len(texto)} totais]")


def _tipo_binario(amostra: bytes, caminho: str) -> str | None:
    """Detecta binário por assinatura, NUL ou alta proporção de controles.

    Um único NUL é suficiente. Bytes >= 128 não contam como controles para que
    texto UTF-8 (inclusive acentuado) não seja classificado como binário.
    """
    for assinatura, tipo in _ASSINATURAS_BINARIAS:
        if amostra.startswith(assinatura):
            return tipo
    if b"\x00" not in amostra and amostra:
        controles = sum(byte in _BYTES_CONTROLE for byte in amostra)
        if controles / len(amostra) <= 0.30:
            return None
    elif b"\x00" not in amostra:
        return None

    extensao = os.path.splitext(caminho)[1].lstrip(".")
    return extensao.upper() if extensao else "desconhecido"


def ler_arquivo(caminho: str, inicio: int = None, fim: int = None) -> str:
    """Lê um arquivo do PROJETO (config.REPO). Opcional: intervalo de linhas
    [inicio, fim] (1-based). Retorna com números de linha p/ facilitar edições."""
    alvo = _dentro(config.REPO, caminho)
    if not os.path.isfile(alvo):
        return f"ERRO: arquivo não existe: {caminho}"
    tamanho = os.path.getsize(alvo)
    with open(alvo, "rb") as f:
        amostra = f.read(8192)
    tipo = _tipo_binario(amostra, caminho)
    if tipo:
        return (f"ERRO: arquivo binário ({tamanho} bytes, tipo detectado: "
                f"{tipo}). Use ferramentas específicas se precisar do conteúdo.")
    with open(alvo, "r", encoding="utf-8", errors="replace") as f:
        linhas = f.readlines()
    total = len(linhas)
    ini = max(1, int(inicio or 1))
    f_ = min(total, int(fim or total)) if total else 0
    largura = len(str(f_)) or 1
    corpo = "".join(f"{ini + i:>{largura}}\t{l}"
                    for i, l in enumerate(linhas[ini - 1:f_]))
    corpo = _truncar(corpo, 20000)
    return f"# {caminho} (linhas {ini}-{f_} de {total})\n{corpo}"


def escrever_arquivo(caminho: str, conteudo: str) -> str:
    """Cria/sobrescreve um arquivo. Aceita caminho relativo (contra o projeto)
    ou absoluto. Caminhos externos exigem confirmação explícita de alto risco."""
    if not permissao.consumir(permissao.comando_de("escrever_arquivo", {"caminho": caminho})):
        return "ERRO: escrita não passou pela aprovação de risco (trinco de segurança)."
    alvo = _resolver_alvo(caminho)
    pasta = os.path.dirname(alvo)
    if pasta:
        os.makedirs(pasta, exist_ok=True)
    with open(alvo, "w", encoding="utf-8") as f:
        f.write(conteudo)
    return f"OK: {len(conteudo)} caracteres gravados em {alvo}"


def listar_diretorio(caminho: str = ".", recursivo: bool = False) -> str:
    """Lista o diretório do PROJETO. recursivo=True mostra a árvore (ignora
    .git/node_modules/.venv/etc.)."""
    alvo = _dentro(config.REPO, caminho)
    if not os.path.isdir(alvo):
        return f"ERRO: diretório não existe: {caminho}"
    if not recursivo:
        itens = [n + ("/" if os.path.isdir(os.path.join(alvo, n)) else "")
                 for n in sorted(os.listdir(alvo))]
        return "\n".join(itens) if itens else "(vazio)"
    linhas = []
    for raiz, dirs, arqs in os.walk(alvo):
        dirs[:] = sorted(d for d in dirs if d not in _IGNORAR)
        rel = os.path.relpath(raiz, alvo)
        nivel = 0 if rel == "." else rel.count(os.sep) + 1
        if rel != ".":
            linhas.append("  " * (nivel - 1) + os.path.basename(raiz) + "/")
        for a in sorted(arqs):
            linhas.append("  " * nivel + a)
        if len(linhas) > 400:
            linhas.append("...[truncado — projeto grande]")
            break
    return "\n".join(linhas) if linhas else "(vazio)"


def buscar_codigo(padrao: str, caminho: str = ".", ext: str = None) -> str:
    """Procura texto/regex nos arquivos do PROJETO (tipo grep -rn). `ext` filtra
    a extensão (ex: '.py'). Retorna 'arquivo:linha: trecho'."""
    import re as _re
    base = _dentro(config.REPO, caminho)
    if not os.path.exists(base):
        return f"ERRO: caminho não existe: {caminho}"
    try:
        rx = _re.compile(padrao)
    except _re.error as e:
        return f"ERRO: regex inválida: {e}"

    def varrer(fp):
        try:
            with open(fp, "r", encoding="utf-8", errors="ignore") as f:
                for n, linha in enumerate(f, 1):
                    if rx.search(linha):
                        rel = os.path.relpath(fp, config.REPO)
                        yield f"{rel}:{n}: {linha.strip()[:200]}"
        except OSError:
            return

    def arquivos():
        if os.path.isfile(base):
            yield base
            return
        for raiz, dirs, arqs in os.walk(base):
            dirs[:] = [d for d in dirs if d not in _IGNORAR]
            for a in sorted(arqs):
                yield os.path.join(raiz, a)

    limite_exibicao = 100
    limite_contagem = 500
    achados = []
    total = 0
    contagem_incompleta = False
    for fp in arquivos():
        if ext and not fp.endswith(ext):
            continue
        for hit in varrer(fp):
            total += 1
            if len(achados) < limite_exibicao:
                achados.append(hit)
            if total > limite_contagem:
                contagem_incompleta = True
                break
        if contagem_incompleta:
            break
    if total > limite_exibicao:
        total_txt = f"{limite_contagem}+" if contagem_incompleta else str(total)
        omitidos_txt = (f"ao menos {limite_contagem + 1 - limite_exibicao}"
                        if contagem_incompleta else str(total - limite_exibicao))
        achados.append(
            f"...[truncado: {omitidos_txt} resultados omitidos; "
            f"{total_txt} resultados no total — refine a busca]"
        )
    return "\n".join(achados) if achados else f"Nada encontrado para: {padrao}"


def rodar_comando(comando: str) -> str:
    # shell=True é necessário para pipes, redirecionamentos, glob e encadeamento.
    comando = comando.strip()
    if not comando:
        return "ERRO: comando vazio"
    if not permissao.consumir(comando):
        return ("ERRO: comando não passou pela aprovação de risco (trinco de "
                "segurança). Chame pelo fluxo normal — não há como pular o gate.")
    try:
        r = subprocess.run(comando, cwd=config.REPO, shell=True,
                           capture_output=True, text=True,
                           timeout=config.TIMEOUT_COMANDO)
    except subprocess.TimeoutExpired:
        return f"ERRO: comando estourou {config.TIMEOUT_COMANDO}s"
    saida = (r.stdout + r.stderr).strip()
    saida = _truncar(saida, 8000) if saida else "(sem saída)"
    return f"{saida}\n[código de saída: {r.returncode}]"


def git(args: str = "") -> str:
    """Roda ``git <args>`` no repositório atual após aprovação."""
    if not permissao.consumir(("git " + (args or "")).strip()):
        return ("ERRO: git não passou pela aprovação de risco (trinco de "
                "segurança). Chame pelo fluxo normal.")
    try:
        partes = shlex.split(args)
    except ValueError as e:
        return f"ERRO: argumentos git inválidos: {e}"
    if partes and partes[0].startswith("-"):
        return ("ERRO: opções globais do git (ex: -c, -C, --exec-path) não são "
                "permitidas antes do subcomando (vetor de injeção). "
                "Use a forma 'git <subcomando> ...'.")
    if not os.path.isdir(os.path.join(config.REPO, ".git")):
        return (f"ERRO: {config.REPO} não é um repositório git "
                f"(rode o hrx dentro de um projeto, ou use 'init').")
    try:
        r = subprocess.run(["git", *partes], cwd=config.REPO, capture_output=True,
                           text=True, timeout=config.TIMEOUT_COMANDO)
    except FileNotFoundError:
        return "ERRO: git não está instalado."
    except subprocess.TimeoutExpired:
        return f"ERRO: git estourou {config.TIMEOUT_COMANDO}s"
    saida = (r.stdout + r.stderr).strip()
    saida = _truncar(saida, 8000) if saida else "(sem saída)"
    return f"{saida}\n[código de saída: {r.returncode}]"


def editar_arquivo(caminho: str, procurar: str, substituir: str,
                   ocorrencia: int = None, tudo: bool = False) -> str:
    """Substitui um trecho literal inequívoco, uma ocorrência ou todas.

    O arquivo pode ser relativo ao projeto ou absoluto. A operação passa pelo
    trinco de aprovação e só escreve depois de validar todos os argumentos.
    """
    if not permissao.consumir(permissao.comando_de("editar_arquivo", {"caminho": caminho})):
        return "ERRO: edição não passou pela aprovação de risco (trinco de segurança)."
    if not isinstance(procurar, str) or not procurar:
        return "ERRO: 'procurar' deve ser uma string não vazia"
    if not isinstance(substituir, str):
        return "ERRO: 'substituir' deve ser uma string"
    if not isinstance(tudo, bool):
        return "ERRO: 'tudo' deve ser booleano (true ou false)"
    if tudo and ocorrencia is not None:
        return "ERRO: use 'ocorrencia=N' ou 'tudo=True', não ambos"
    if (ocorrencia is not None
            and (isinstance(ocorrencia, bool) or not isinstance(ocorrencia, int))):
        return "ERRO: 'ocorrencia' deve ser um inteiro a partir de 1"
    if ocorrencia is not None and ocorrencia < 1:
        return "ERRO: 'ocorrencia' deve ser um inteiro a partir de 1"

    alvo = _resolver_alvo(caminho)
    if not os.path.isfile(alvo):
        return f"ERRO: arquivo não existe: {alvo}"
    with open(alvo, "r", encoding="utf-8", errors="replace") as f:
        texto = f.read()
    if procurar not in texto:
        return f"ERRO: trecho a procurar não encontrado em {alvo}"
    n = texto.count(procurar)
    if not tudo and ocorrencia is None and n > 1:
        return (f"ERRO: 'procurar' aparece {n}x em {alvo}. Passe "
                f"`ocorrencia=N` (1..{n}) ou `tudo=True`, ou aumente o "
                "contexto de `procurar` para deixá-lo único.")
    if ocorrencia is not None and ocorrencia > n:
        return (f"ERRO: 'ocorrencia={ocorrencia}' está fora da faixa; "
                f"'procurar' aparece {n}x em {alvo}. Use um valor de 1 a {n}.")

    if tudo:
        novo_texto = texto.replace(procurar, substituir)
        substituidas = n
    else:
        escolhida = ocorrencia or 1
        inicio = 0
        indice = -1
        for _ in range(escolhida):
            indice = texto.find(procurar, inicio)
            inicio = indice + len(procurar)
        novo_texto = texto[:indice] + substituir + texto[indice + len(procurar):]
        substituidas = 1

    with open(alvo, "w", encoding="utf-8") as f:
        f.write(novo_texto)
    return f"OK: {substituidas} ocorrência(s) substituída(s) em {alvo}"


_HUNK_PATCH = re.compile(
    r"^@@ -(\d+)(?:,(\d+))? \+(\d+)(?:,(\d+))? @@(?: .*)?$"
)


def aplicar_patch(caminho: str, patch: str) -> str:
    """Aplica hunks de um diff unificado a um arquivo de forma atômica."""
    comando = permissao.comando_de("aplicar_patch", {"caminho": caminho})
    if not permissao.consumir(comando):
        return "ERRO: patch não passou pela aprovação de risco (trinco de segurança)."
    alvo = _resolver_alvo(caminho)
    if not os.path.isfile(alvo):
        return f"ERRO: arquivo não existe: {alvo}"
    with open(alvo, "r", encoding="utf-8", errors="replace", newline="") as f:
        texto = f.read()

    origem = texto.splitlines()
    linhas_patch = str(patch or "").splitlines()
    hunks = []
    atual = None
    saida_sem_nova_linha = not texto.endswith(("\n", "\r"))
    antiga_sem_nova_linha = False
    ultimo_marcador = None
    for numero, linha in enumerate(linhas_patch, 1):
        cabecalho = _HUNK_PATCH.match(linha)
        if cabecalho:
            if atual is not None:
                hunks.append(atual)
            antigo_inicio = int(cabecalho.group(1))
            antigo_total = int(cabecalho.group(2) or 1)
            novo_total = int(cabecalho.group(4) or 1)
            atual = {
                "linha": numero,
                "inicio": antigo_inicio - 1 if antigo_total else antigo_inicio,
                "antigo_total": antigo_total,
                "novo_total": novo_total,
                "antigas": [],
                "novas": [],
            }
            continue
        if atual is None:
            if not linha or linha.startswith(("--- ", "+++ ")):
                continue
            return f"ERRO: patch inválido na linha {numero}: esperado cabeçalho @@"
        if linha.startswith("\\ No newline at end of file"):
            if ultimo_marcador in "+ ":
                saida_sem_nova_linha = True
            elif ultimo_marcador == "-":
                antiga_sem_nova_linha = True
            continue
        if not linha or linha[0] not in " +-":
            return f"ERRO: patch inválido na linha {numero}: marcador desconhecido"
        marcador, conteudo = linha[0], linha[1:]
        if marcador == "+" and antiga_sem_nova_linha:
            saida_sem_nova_linha = False
            antiga_sem_nova_linha = False
        if marcador in " -":
            atual["antigas"].append(conteudo)
        if marcador in " +":
            atual["novas"].append(conteudo)
        ultimo_marcador = marcador
    if atual is not None:
        hunks.append(atual)
    if not hunks:
        return "ERRO: patch não contém nenhum hunk @@"

    resultado = []
    cursor = 0
    for hunk in hunks:
        inicio = hunk["inicio"]
        antigas = hunk["antigas"]
        novas = hunk["novas"]
        if len(antigas) != hunk["antigo_total"] or len(novas) != hunk["novo_total"]:
            return f"ERRO: contagem inválida no hunk da linha {hunk['linha']}"
        if inicio < cursor or inicio > len(origem):
            return f"ERRO: posição inválida no hunk da linha {hunk['linha']}"
        if origem[inicio:inicio + len(antigas)] != antigas:
            return f"ERRO: conflito no hunk da linha {hunk['linha']}; arquivo não alterado"
        resultado.extend(origem[cursor:inicio])
        resultado.extend(novas)
        cursor = inicio + len(antigas)
    resultado.extend(origem[cursor:])

    quebra = "\r\n" if "\r\n" in texto else "\r" if "\r" in texto else "\n"
    novo_texto = quebra.join(resultado)
    if resultado and not saida_sem_nova_linha:
        novo_texto += quebra
    temporario = None
    try:
        with tempfile.NamedTemporaryFile(
            "w", encoding="utf-8", newline="", dir=os.path.dirname(alvo),
            prefix=".hrx-patch-", delete=False,
        ) as f:
            temporario = f.name
            f.write(novo_texto)
        os.chmod(temporario, os.stat(alvo).st_mode & 0o777)
        os.replace(temporario, alvo)
    finally:
        if temporario and os.path.exists(temporario):
            os.unlink(temporario)
    return f"OK: {len(hunks)} hunk(s) aplicado(s) em {alvo}"


def criar_planilha(nome: str, dados: list, cabecalho: list = None) -> str:
    """Cria uma planilha Excel (.xlsx). `nome` pode ser relativo ao projeto ou
    absoluto (`~/Downloads/x.xlsx`, `/tmp/y.xlsx`). `dados` = lista de linhas
    (listas) ou lista de dicionários. `cabecalho` opcional (lista de colunas)."""
    if not permissao.consumir(permissao.comando_de("criar_planilha", {"nome": nome})):
        return "ERRO: criação de planilha não passou pela aprovação de risco (trinco de segurança)."
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if not nome.lower().endswith(".xlsx"):
        nome += ".xlsx"
    alvo = _resolver_alvo(nome)
    pasta = os.path.dirname(alvo)
    if pasta:
        os.makedirs(pasta, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    linhas = list(dados or [])
    tem_cabecalho = False

    if linhas and isinstance(linhas[0], dict):
        cols = cabecalho or list(linhas[0].keys())
        ws.append(cols)
        tem_cabecalho = True
        for d in linhas:
            ws.append([d.get(c, "") for c in cols])
    else:
        if cabecalho:
            ws.append(list(cabecalho))
            tem_cabecalho = True
        for row in linhas:
            ws.append(list(row) if isinstance(row, (list, tuple)) else [row])

    if tem_cabecalho:
        for c in ws[1]:
            c.font = Font(bold=True)
    for col in ws.columns:
        largura = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(largura + 2, 60)

    wb.save(alvo)
    return f"OK: planilha criada em {alvo} ({ws.max_row} linhas x {ws.max_column} colunas)"


def criar_pdf(nome: str, titulo: str = None, conteudo=None, tabela: list = None) -> str:
    """Cria um PDF. `nome` pode ser relativo ao projeto ou absoluto
    (`~/Downloads/x.pdf`, `/tmp/y.pdf`). `titulo` (opcional), `conteudo` (texto
    ou lista de parágrafos), `tabela` (opcional: lista de linhas, a 1ª vira
    cabeçalho)."""
    if not permissao.consumir(permissao.comando_de("criar_pdf", {"nome": nome})):
        return "ERRO: criação de PDF não passou pela aprovação de risco (trinco de segurança)."
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    if not nome.lower().endswith(".pdf"):
        nome += ".pdf"
    alvo = _resolver_alvo(nome)
    pasta = os.path.dirname(alvo)
    if pasta:
        os.makedirs(pasta, exist_ok=True)

    estilos = getSampleStyleSheet()
    flow = []
    if titulo:
        flow.append(Paragraph(str(titulo), estilos["Title"]))
        flow.append(Spacer(1, 12))
    if conteudo:
        paras = conteudo if isinstance(conteudo, list) else [conteudo]
        for p in paras:
            flow.append(Paragraph(str(p), estilos["BodyText"]))
            flow.append(Spacer(1, 6))
    if tabela:
        t = Table([[str(c) for c in linha] for linha in tabela], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8b0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        flow.append(Spacer(1, 8))
        flow.append(t)

    SimpleDocTemplate(alvo, pagesize=A4).build(flow)
    return f"OK: PDF criado em {alvo}"


def _fmt_cve(cve: dict) -> str:
    cid = cve.get("id", "?")
    desc = next((d["value"] for d in cve.get("descriptions", [])
                 if d.get("lang") == "en"), "(sem descrição)")
    metrics = cve.get("metrics", {})
    score = severidade = vetor = None
    for chave in ("cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        if metrics.get(chave):
            d = metrics[chave][0].get("cvssData", {})
            score = d.get("baseScore")
            severidade = d.get("baseSeverity") or metrics[chave][0].get("baseSeverity")
            vetor = d.get("vectorString")
            break
    refs = [r.get("url") for r in cve.get("references", [])][:3]
    linhas = [f"🔹 {cid}"]
    if score is not None:
        linhas.append(f"   CVSS: {score} ({severidade or '?'})  {vetor or ''}".rstrip())
    if cve.get("published"):
        linhas.append(f"   publicado: {cve['published'][:10]}")
    linhas.append(f"   {desc[:600]}")
    for u in refs:
        linhas.append(f"   ref: {u}")
    return "\n".join(linhas)


def consultar_cve(consulta: str) -> str:
    """Consulta CVEs na API oficial do NVD (NIST). Aceita um ID (CVE-2021-44228)
    ou uma palavra-chave (ex: 'log4j', 'openssl heartbleed')."""
    import re
    import requests

    consulta = (consulta or "").strip()
    if not consulta:
        return "ERRO: informe um CVE-ID ou palavra-chave"

    base = "https://services.nvd.nist.gov/rest/json/cves/2.0"
    m = re.search(r"CVE-\d{4}-\d{4,7}", consulta, re.IGNORECASE)
    params = ({"cveId": m.group(0).upper()} if m
              else {"keywordSearch": consulta, "resultsPerPage": 5})
    try:
        r = requests.get(base, params=params,
                         headers={"User-Agent": f"{config.NOME}-agente"}, timeout=20)
    except requests.RequestException as e:
        return f"ERRO ao consultar o NVD (sem internet?): {e}"
    if r.status_code == 404:
        return f"Nada encontrado para: {consulta}"
    if r.status_code != 200:
        return f"ERRO: NVD respondeu {r.status_code} (limite de requisições? tente de novo em ~30s)"

    vulns = r.json().get("vulnerabilities", [])
    if not vulns:
        return f"Nada encontrado para: {consulta}"
    blocos = [_fmt_cve(v.get("cve", {})) for v in vulns[:5]]
    return "\n\n".join(blocos)


def _ler_memoria() -> list:
    if not os.path.isfile(config.MEMORIA):
        return []
    try:
        with open(config.MEMORIA, "r", encoding="utf-8") as f:
            dados = json.load(f)
        return dados if isinstance(dados, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def _gravar_memoria(itens: list) -> None:
    _garantir()
    with open(config.MEMORIA, "w", encoding="utf-8") as f:
        json.dump(itens, f, ensure_ascii=False, indent=2)


def _gravar_memoria_raw(itens: list) -> None:
    _garantir()
    with open(config.MEMORIA, "w", encoding="utf-8") as f:
        json.dump(itens, f, ensure_ascii=False, indent=2)


def _gravar_resumo(texto: str) -> None:
    _garantir()
    with open(_arquivo_resumo(), "w", encoding="utf-8") as f:
        f.write((texto or "").strip())


def _ler_resumo() -> str:
    try:
        with open(_arquivo_resumo(), "r", encoding="utf-8") as f:
            return f.read().strip()
    except OSError:
        return ""


def _relevancia_memoria(item: dict, consulta: str = "") -> int:
    """Pontua memórias pelo tipo, recência e sobreposição com a conversa atual."""
    tipo = str(item.get("tipo", "fato")).strip().lower()
    pesos = {
        "decisao": 120,
        "projeto": 100,
        "preferencia": 90,
        "comando": 80,
        "fato": 60,
        "resumo": 50,
    }
    score = pesos.get(tipo, 40)
    texto = str(item.get("texto", "")).lower()
    consulta = (consulta or "").lower()
    for termo in consulta.split():
        if len(termo) > 3 and termo in texto:
            score += 8
    try:
        score += min(20, int(item.get("id", 0)) // 3)
    except (TypeError, ValueError):
        pass
    return score


def _resumir_memorias(itens: list) -> str:
    selecionadas = itens[:max(1, int(config.MEMORIA_PROMPT_RESUMO_ITENS))]
    linhas = []
    total = 0
    for m in selecionadas:
        texto = " ".join(str(m.get("texto", "")).strip().split())
        if len(texto) > 160:
            texto = texto[:157] + "..."
        linha = f"- [{m.get('tipo', 'fato')}] {texto}"
        if total + len(linha) > config.MEMORIA_PROMPT_RESUMO_CHARS and linhas:
            break
        linhas.append(linha)
        total += len(linha)
    return "\n".join(linhas)


def _compactar_memoria() -> None:
    itens = _ler_memoria()
    if len(itens) <= max(1, int(config.MEMORIA_PROMPT_RESUMO_A_PARTIR)):
        return
    itens = sorted(itens, key=lambda m: int(m.get("id", 0)) or 0)
    manter = max(1, int(config.MEMORIA_PROMPT_MAX_ITENS))
    antigos = itens[:-manter]
    novos = itens[-manter:]
    resumo_atual = _ler_resumo()
    partes = [p for p in [resumo_atual, _resumir_memorias(antigos)] if p]
    resumo = "\n\n".join(partes).strip()
    if resumo:
        _gravar_resumo(resumo)
    _gravar_memoria_raw(novos)


def carregar_memorias() -> list:
    """Lista de memórias, usada pelo agente para injetar no contexto."""
    memorias = _ler_memoria()
    resumo = _ler_resumo()
    if resumo:
        memorias = [{"id": 0, "tipo": "resumo", "texto": resumo}] + memorias
    return memorias


def _prioridade_memoria(item: dict) -> tuple:
    """Ordena memórias por utilidade provável no prompt."""
    tipo = str(item.get("tipo", "fato")).strip().lower()
    pesos = {
        "decisao": 0,
        "projeto": 1,
        "comando": 2,
        "preferencia": 3,
        "fato": 4,
    }
    return (pesos.get(tipo, 9), -(int(item.get("id", 0)) or 0))


def memoria_contexto(max_itens: int = None, max_chars: int = None, consulta: str = "") -> str:
    """Versão compacta da memória para uso no prompt.
    Mantém poucas memórias e corta textos longos para economizar contexto."""
    itens = carregar_memorias()
    if not itens:
        return "(nenhuma memória guardada)"
    if str(getattr(config, "MEMORIA_PROMPT", "compacta")).strip().lower() == "completa":
        return "\n".join(
            f"- #{m.get('id', '?')} [{m.get('tipo', 'fato')}] "
            f"{' '.join(str(m.get('texto', '')).strip().split())}"
            for m in sorted(itens, key=lambda x: (_prioridade_memoria(x), -len(str(x.get("texto", "")))))
        )
    limite_itens = max_itens if max_itens is not None else config.MEMORIA_PROMPT_MAX_ITENS
    limite_chars = max_chars if max_chars is not None else config.MEMORIA_PROMPT_MAX_CHARS
    contexto = consulta or " ".join(str(m.get("texto", "")) for m in itens[-3:])
    ordenadas = sorted(
        itens,
        key=lambda m: (_relevancia_memoria(m, contexto), _prioridade_memoria(m)),
        reverse=True,
    )
    selecionadas = ordenadas[:max(1, int(limite_itens))]
    linhas = []
    total = 0
    for m in selecionadas:
        texto = str(m.get("texto", "")).strip().replace("\n", " ")
        texto = " ".join(texto.split())
        if len(texto) > 120:
            texto = texto[:117] + "..."
        linha = f"- #{m.get('id', '?')} [{m.get('tipo', 'fato')}] {texto}"
        if total + len(linha) > limite_chars and linhas:
            break
        linhas.append(linha)
        total += len(linha)
    return "\n".join(linhas) if linhas else "(nenhuma memória guardada)"


def memoria_salvar(texto: str, tipo: str = "fato") -> str:
    """Guarda um fato/decisão/comando para lembrar em sessões futuras.
    tipo sugerido: fato | comando | decisao | projeto."""
    texto = (texto or "").strip()
    if not texto:
        return "ERRO: nada para lembrar (texto vazio)"
    itens = _ler_memoria()
    if any(m.get("texto", "").strip().lower() == texto.lower() for m in itens):
        return "OK: já estava na memória (sem duplicar)."
    novo_id = max((m.get("id", 0) for m in itens), default=0) + 1
    itens.append({"id": novo_id, "ts": datetime.date.today().isoformat(),
                  "tipo": (tipo or "fato").strip(), "texto": texto})
    _gravar_memoria(itens)
    _compactar_memoria()
    return f"OK: memória #{novo_id} guardada ({tipo})."


def memoria_listar() -> str:
    """Lista tudo que o agente já guardou na memória."""
    itens = carregar_memorias()
    if not itens:
        return "(nenhuma memória guardada)"
    return "\n".join(f"#{m['id']} [{m.get('tipo', 'fato')}] {m['texto']}"
                     for m in itens)


def memoria_contexto_compacto() -> str:
    return memoria_contexto()


def memoria_resumir() -> str:
    _compactar_memoria()
    resumo = _ler_resumo()
    return resumo or "(nenhum resumo disponível)"


def memoria_limpar() -> str:
    _garantir()
    for arq in (config.MEMORIA, _arquivo_resumo()):
        try:
            os.remove(arq)
        except OSError:
            pass
    return "OK: memória e resumo limpos."


def memoria_esquecer(alvo: str) -> str:
    """Remove memórias por #id (ex: '3') ou por termo contido no texto."""
    alvo = str(alvo or "").strip()
    if not alvo:
        return "ERRO: informe o #id ou um termo da memória a esquecer"
    itens = _ler_memoria()
    antes = len(itens)
    if alvo.lstrip("#").isdigit():
        idx = int(alvo.lstrip("#"))
        itens = [m for m in itens if m.get("id") != idx]
    else:
        termo = alvo.lower()
        itens = [m for m in itens if termo not in m.get("texto", "").lower()]
    removidas = antes - len(itens)
    if removidas:
        _gravar_memoria(itens)
        return f"OK: {removidas} memória(s) esquecida(s)."
    return "Nenhuma memória correspondeu."


def buscar_docs(consulta: str) -> str:
    _garantir()
    termos = [t.lower() for t in consulta.split() if len(t) > 2]
    if not termos:
        return "ERRO: consulta muito curta"
    achados = []
    for arq in glob.glob(os.path.join(config.DADOS, "**", "*"), recursive=True):
        if not os.path.isfile(arq):
            continue
        if os.path.splitext(arq)[1].lower() not in (".txt", ".md", ".csv", ".json", ""):
            continue
        try:
            texto = open(arq, "r", encoding="utf-8", errors="replace").read()
        except OSError:
            continue
        for i in range(0, len(texto), 500):
            bloco = texto[i:i + 700]
            score = sum(bloco.lower().count(t) for t in termos)
            if score:
                achados.append((score, os.path.basename(arq), bloco.strip()))
    achados.sort(key=lambda x: x[0], reverse=True)
    if not achados:
        return "Nada encontrado nos documentos."
    return "\n\n---\n\n".join(f"[{n}] (rel. {s})\n{b[:500]}" for s, n, b in achados[:4])


_WEB_TAGS_IGNORAR = {"script", "style", "noscript", "svg", "iframe",
                     "nav", "header", "footer", "aside", "form", "template"}
_WEB_TAGS_BLOCO = {"p", "div", "section", "article", "li", "tr",
                   "table", "ul", "ol", "blockquote", "br", "hr"}
_WEB_TAGS_H = {"h1": "# ", "h2": "## ", "h3": "### ",
               "h4": "#### ", "h5": "##### ", "h6": "###### "}


def _html_para_texto(html: str) -> str:
    """Converte HTML em texto próximo de Markdown."""
    import re as _re
    from html.parser import HTMLParser

    partes = []
    ignorar = [0]
    href = [None]

    class _P(HTMLParser):
        def handle_starttag(self, tag, attrs):
            if tag in _WEB_TAGS_IGNORAR:
                ignorar[0] += 1
                return
            if ignorar[0]:
                return
            if tag in _WEB_TAGS_H:
                partes.append("\n\n" + _WEB_TAGS_H[tag])
            elif tag == "pre":
                partes.append("\n\n```\n")
            elif tag == "code":
                partes.append("`")
            elif tag == "a":
                href[0] = dict(attrs).get("href", "")
            elif tag in _WEB_TAGS_BLOCO:
                partes.append("\n")

        def handle_endtag(self, tag):
            if tag in _WEB_TAGS_IGNORAR:
                ignorar[0] = max(0, ignorar[0] - 1)
                return
            if ignorar[0]:
                return
            if tag in _WEB_TAGS_H or tag in _WEB_TAGS_BLOCO:
                partes.append("\n")
            elif tag == "pre":
                partes.append("\n```\n")
            elif tag == "code":
                partes.append("`")
            elif tag == "a" and href[0]:
                partes.append(f" [{href[0]}]")
                href[0] = None

        def handle_data(self, data):
            if not ignorar[0]:
                partes.append(data)

    try:
        _P().feed(html)
    except Exception:  # noqa: BLE001 — HTMLParser é tolerante, mas garante fallback
        return _re.sub(r"<[^>]+>", " ", html)
    texto = "".join(partes)
    texto = _re.sub(r"[ \t]+", " ", texto)
    texto = _re.sub(r"\n{3,}", "\n\n", texto)
    return texto.strip()


def _validar_url_publica(url: str) -> str:
    """Valida o esquema e bloqueia endereços internos contra SSRF.

    Ainda há risco de DNS rebinding entre a validação e a conexão; eliminá-lo
    exige fixar o IP e o cabeçalho Host em um adaptador HTTP próprio.
    """
    import ipaddress
    import socket
    from urllib.parse import urlparse

    p = urlparse(url)
    if p.scheme not in ("http", "https"):
        raise ValueError(f"esquema não permitido: {p.scheme or '(vazio)'}")
    host = p.hostname
    if not host:
        raise ValueError("URL sem host")
    try:
        infos = socket.getaddrinfo(host, None)
    except socket.gaierror:
        raise ValueError(f"host não resolve: {host}")
    if not infos:
        raise ValueError(f"host sem endereços: {host}")
    for info in infos:
        endereco = info[4][0]
        ip = ipaddress.ip_address(endereco)
        # Normaliza IPv4 mapeado em IPv6 antes da validação.
        if isinstance(ip, ipaddress.IPv6Address) and ip.ipv4_mapped:
            ip = ip.ipv4_mapped
        if (ip.is_private or ip.is_loopback or ip.is_link_local
                or ip.is_reserved or ip.is_multicast or ip.is_unspecified):
            raise ValueError(f"IP interno bloqueado: {endereco} ({host})")
    return url


def buscar_web(url: str, max_chars: int = 8000) -> str:
    """Baixa uma URL pública e devolve o texto principal em ~markdown.
    Uso: consultar docs, ler stack traces, ver issues/CVE, blogs técnicos.
    Bloqueia URLs internas (SSRF), esquemas não-http e conteúdo > 2 MB."""
    import requests
    from urllib.parse import urljoin

    if not url or not str(url).strip():
        return "ERRO: URL vazia"
    try:
        _validar_url_publica(url)
    except ValueError as e:
        return f"ERRO: {e}"

    headers = {"User-Agent": "hrx-code/1.0 (+https://www.dwsolutions.company)"}
    atual = url
    for _ in range(6):  # até 5 redirects + a request final
        try:
            r = requests.get(atual, headers=headers, timeout=15,
                             allow_redirects=False, stream=True)
        except requests.RequestException as e:
            return f"ERRO ao baixar {atual}: {e}"
        if r.is_redirect or r.status_code in (301, 302, 303, 307, 308):
            destino = r.headers.get("Location", "")
            r.close()
            if not destino:
                return f"ERRO: redirect sem Location em {atual}"
            atual = urljoin(atual, destino)
            try:
                _validar_url_publica(atual)
            except ValueError as e:
                return f"ERRO ao seguir redirect: {e}"
            continue
        break
    else:
        return "ERRO: muitos redirects (>5)"

    if not r.ok:
        r.close()
        return f"ERRO HTTP {r.status_code} em {atual}"

    limite_bytes = 2_000_000
    corpo = bytearray()
    for chunk in r.iter_content(65536):
        corpo += chunk
        if len(corpo) >= limite_bytes:
            corpo = corpo[:limite_bytes]
            break
    ctype = (r.headers.get("Content-Type") or "").lower()
    encoding = r.encoding or "utf-8"
    r.close()

    try:
        texto = bytes(corpo).decode(encoding, errors="replace")
    except LookupError:
        texto = bytes(corpo).decode("utf-8", errors="replace")

    inicio = texto.lstrip()[:15].lower()
    parece_html = ("html" in ctype or inicio.startswith("<!doctype")
                   or inicio.startswith("<html"))
    if parece_html:
        texto = _html_para_texto(texto)

    if len(texto) > max_chars:
        texto = texto[:max_chars] + f"\n\n...[truncado — ~{len(corpo)} bytes baixados]"
    return f"# {atual}\n\n{texto}"


REGISTRO = {
    "ler_arquivo": ler_arquivo,
    "escrever_arquivo": escrever_arquivo,
    "editar_arquivo": editar_arquivo,
    "aplicar_patch": aplicar_patch,
    "listar_diretorio": listar_diretorio,
    "buscar_codigo": buscar_codigo,
    "criar_planilha": criar_planilha,
    "criar_pdf": criar_pdf,
    "rodar_comando": rodar_comando,
    "git": git,
    "consultar_cve": consultar_cve,
    "buscar_web": buscar_web,
    "memoria_salvar": memoria_salvar,
    "memoria_listar": memoria_listar,
    "memoria_esquecer": memoria_esquecer,
    "memoria_resumir": memoria_resumir,
    "memoria_limpar": memoria_limpar,
    "buscar_docs": buscar_docs,
}


def executar(nome: str, args: dict) -> str:
    fn = REGISTRO.get(nome)
    if fn is None:
        return f"ERRO: ferramenta desconhecida '{nome}'"
    try:
        return fn(**(args or {}))
    except TypeError as e:
        return f"ERRO: argumentos inválidos para {nome}: {e}"
    except Exception as e:  # noqa: BLE001
        return f"ERRO ao executar {nome}: {e}"
