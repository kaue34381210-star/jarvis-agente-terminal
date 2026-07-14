"""Ferramentas do agente. Arquivos restritos a WORKSPACE/ e DADOS/;
o git opera no repositório do diretório atual (config.REPO)."""
import os
import glob
import shlex
import subprocess

import config


def _dentro(base: str, caminho: str) -> str:
    alvo = os.path.normpath(os.path.join(base, caminho))
    if alvo != base and not alvo.startswith(base + os.sep):
        raise ValueError(f"Caminho fora da área permitida (bloqueado): {caminho}")
    return alvo


def _garantir() -> None:
    os.makedirs(config.WORKSPACE, exist_ok=True)
    os.makedirs(config.DADOS, exist_ok=True)


def ler_arquivo(caminho: str) -> str:
    alvo = _dentro(config.WORKSPACE, caminho)
    if not os.path.isfile(alvo):
        return f"ERRO: arquivo não existe: {caminho}"
    with open(alvo, "r", encoding="utf-8", errors="replace") as f:
        conteudo = f.read()
    return conteudo[:20000] + ("\n...[truncado]" if len(conteudo) > 20000 else "")


def escrever_arquivo(caminho: str, conteudo: str) -> str:
    alvo = _dentro(config.WORKSPACE, caminho)
    os.makedirs(os.path.dirname(alvo) or config.WORKSPACE, exist_ok=True)
    with open(alvo, "w", encoding="utf-8") as f:
        f.write(conteudo)
    return f"OK: {len(conteudo)} caracteres gravados em {caminho}"


def listar_diretorio(caminho: str = ".") -> str:
    alvo = _dentro(config.WORKSPACE, caminho)
    if not os.path.isdir(alvo):
        return f"ERRO: diretório não existe: {caminho}"
    itens = [n + ("/" if os.path.isdir(os.path.join(alvo, n)) else "")
             for n in sorted(os.listdir(alvo))]
    return "\n".join(itens) if itens else "(vazio)"


def rodar_comando(comando: str) -> str:
    # A liberação por risco (🟢🟡🔴) é feita na camada de aprovação (agente.py).
    # Aqui só executamos com sandbox de diretório e timeout.
    partes = comando.strip().split()
    if not partes:
        return "ERRO: comando vazio"
    try:
        r = subprocess.run(partes, cwd=config.WORKSPACE, capture_output=True,
                           text=True, timeout=config.TIMEOUT_COMANDO)
    except subprocess.TimeoutExpired:
        return f"ERRO: comando estourou {config.TIMEOUT_COMANDO}s"
    saida = (r.stdout + r.stderr).strip()
    return saida[:8000] if saida else f"(sem saída, código {r.returncode})"


def git(args: str = "") -> str:
    """Roda `git <args>` no repositório atual (config.REPO). A liberação por
    risco (🟢🟡🔴) acontece na camada de aprovação (agente.py)."""
    try:
        partes = shlex.split(args)
    except ValueError as e:
        return f"ERRO: argumentos git inválidos: {e}"
    if not os.path.isdir(os.path.join(config.REPO, ".git")):
        return (f"ERRO: {config.REPO} não é um repositório git "
                f"(rode o jarvis dentro de um projeto, ou use 'init').")
    try:
        r = subprocess.run(["git", *partes], cwd=config.REPO, capture_output=True,
                           text=True, timeout=config.TIMEOUT_COMANDO)
    except FileNotFoundError:
        return "ERRO: git não está instalado."
    except subprocess.TimeoutExpired:
        return f"ERRO: git estourou {config.TIMEOUT_COMANDO}s"
    saida = (r.stdout + r.stderr).strip()
    return saida[:8000] if saida else f"(sem saída, código {r.returncode})"


def editar_arquivo(caminho: str, procurar: str, substituir: str) -> str:
    """Substitui texto num arquivo existente do workspace (busca-e-substitui)."""
    alvo = _dentro(config.WORKSPACE, caminho)
    if not os.path.isfile(alvo):
        return f"ERRO: arquivo não existe: {caminho}"
    with open(alvo, "r", encoding="utf-8", errors="replace") as f:
        texto = f.read()
    if procurar not in texto:
        return f"ERRO: trecho a procurar não encontrado em {caminho}"
    n = texto.count(procurar)
    with open(alvo, "w", encoding="utf-8") as f:
        f.write(texto.replace(procurar, substituir))
    return f"OK: {n} ocorrência(s) substituída(s) em {caminho}"


def criar_planilha(nome: str, dados: list, cabecalho: list = None) -> str:
    """Cria uma planilha Excel (.xlsx). `dados` = lista de linhas (listas) ou
    lista de dicionários. `cabecalho` opcional (lista de colunas)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if not nome.lower().endswith(".xlsx"):
        nome += ".xlsx"
    alvo = _dentro(config.WORKSPACE, nome)
    os.makedirs(os.path.dirname(alvo) or config.WORKSPACE, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    linhas = list(dados or [])
    tem_cabecalho = False

    if linhas and isinstance(linhas[0], dict):
        cols = cabecalho or list(linhas[0].keys())
        ws.append(cols)
        tem_cabecalho = True
        for d in linhas:
            ws.append([d.get(c, "") for c in cols])
    else:
        if cabecalho:
            ws.append(list(cabecalho))
            tem_cabecalho = True
        for row in linhas:
            ws.append(list(row) if isinstance(row, (list, tuple)) else [row])

    if tem_cabecalho:
        for c in ws[1]:
            c.font = Font(bold=True)
    # largura automática simples
    for col in ws.columns:
        largura = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(largura + 2, 60)

    wb.save(alvo)
    return f"OK: planilha {nome} criada ({ws.max_row} linhas x {ws.max_column} colunas)"


def criar_pdf(nome: str, titulo: str = None, conteudo=None, tabela: list = None) -> str:
    """Cria um PDF. `titulo` (opcional), `conteudo` (texto ou lista de parágrafos),
    `tabela` (opcional: lista de linhas, a 1ª vira cabeçalho)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    if not nome.lower().endswith(".pdf"):
        nome += ".pdf"
    alvo = _dentro(config.WORKSPACE, nome)
    os.makedirs(os.path.dirname(alvo) or config.WORKSPACE, exist_ok=True)

    estilos = getSampleStyleSheet()
    flow = []
    if titulo:
        flow.append(Paragraph(str(titulo), estilos["Title"]))
        flow.append(Spacer(1, 12))
    if conteudo:
        paras = conteudo if isinstance(conteudo, list) else [conteudo]
        for p in paras:
            flow.append(Paragraph(str(p), estilos["BodyText"]))
            flow.append(Spacer(1, 6))
    if tabela:
        t = Table([[str(c) for c in linha] for linha in tabela], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8b0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        flow.append(Spacer(1, 8))
        flow.append(t)

    SimpleDocTemplate(alvo, pagesize=A4).build(flow)
    return f"OK: PDF {nome} criado"


def _fmt_cve(cve: dict) -> str:
    cid = cve.get("id", "?")
    desc = next((d["value"] for d in cve.get("descriptions", [])
                 if d.get("lang") == "en"), "(sem descrição)")
    metrics = cve.get("metrics", {})
    score = severidade = vetor = None
    for chave in ("cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        if metrics.get(chave):
            d = metrics[chave][0].get("cvssData", {})
            score = d.get("baseScore")
            severidade = d.get("baseSeverity") or metrics[chave][0].get("baseSeverity")
            vetor = d.get("vectorString")
            break
    refs = [r.get("url") for r in cve.get("references", [])][:3]
    linhas = [f"🔹 {cid}"]
    if score is not None:
        linhas.append(f"   CVSS: {score} ({severidade or '?'})  {vetor or ''}".rstrip())
    if cve.get("published"):
        linhas.append(f"   publicado: {cve['published'][:10]}")
    linhas.append(f"   {desc[:600]}")
    for u in refs:
        linhas.append(f"   ref: {u}")
    return "\n".join(linhas)


def consultar_cve(consulta: str) -> str:
    """Consulta CVEs na API oficial do NVD (NIST). Aceita um ID (CVE-2021-44228)
    ou uma palavra-chave (ex: 'log4j', 'openssl heartbleed')."""
    import re
    import requests

    consulta = (consulta or "").strip()
    if not consulta:
        return "ERRO: informe um CVE-ID ou palavra-chave"

    base = "https://services.nvd.nist.gov/rest/json/cves/2.0"
    m = re.search(r"CVE-\d{4}-\d{4,7}", consulta, re.IGNORECASE)
    params = ({"cveId": m.group(0).upper()} if m
              else {"keywordSearch": consulta, "resultsPerPage": 5})
    try:
        r = requests.get(base, params=params,
                         headers={"User-Agent": f"{config.NOME}-agente"}, timeout=20)
    except requests.RequestException as e:
        return f"ERRO ao consultar o NVD (sem internet?): {e}"
    if r.status_code == 404:
        return f"Nada encontrado para: {consulta}"
    if r.status_code != 200:
        return f"ERRO: NVD respondeu {r.status_code} (limite de requisições? tente de novo em ~30s)"

    vulns = r.json().get("vulnerabilities", [])
    if not vulns:
        return f"Nada encontrado para: {consulta}"
    blocos = [_fmt_cve(v.get("cve", {})) for v in vulns[:5]]
    return "\n\n".join(blocos)


def buscar_docs(consulta: str) -> str:
    _garantir()
    termos = [t.lower() for t in consulta.split() if len(t) > 2]
    if not termos:
        return "ERRO: consulta muito curta"
    achados = []
    for arq in glob.glob(os.path.join(config.DADOS, "**", "*"), recursive=True):
        if not os.path.isfile(arq):
            continue
        if os.path.splitext(arq)[1].lower() not in (".txt", ".md", ".csv", ".json", ""):
            continue
        try:
            texto = open(arq, "r", encoding="utf-8", errors="replace").read()
        except OSError:
            continue
        for i in range(0, len(texto), 500):
            bloco = texto[i:i + 700]
            score = sum(bloco.lower().count(t) for t in termos)
            if score:
                achados.append((score, os.path.basename(arq), bloco.strip()))
    achados.sort(key=lambda x: x[0], reverse=True)
    if not achados:
        return "Nada encontrado nos documentos."
    return "\n\n---\n\n".join(f"[{n}] (rel. {s})\n{b[:500]}" for s, n, b in achados[:4])


REGISTRO = {
    "ler_arquivo": ler_arquivo,
    "escrever_arquivo": escrever_arquivo,
    "editar_arquivo": editar_arquivo,
    "listar_diretorio": listar_diretorio,
    "criar_planilha": criar_planilha,
    "criar_pdf": criar_pdf,
    "rodar_comando": rodar_comando,
    "git": git,
    "consultar_cve": consultar_cve,
    "buscar_docs": buscar_docs,
}


def executar(nome: str, args: dict) -> str:
    fn = REGISTRO.get(nome)
    if fn is None:
        return f"ERRO: ferramenta desconhecida '{nome}'"
    try:
        return fn(**(args or {}))
    except TypeError as e:
        return f"ERRO: argumentos inválidos para {nome}: {e}"
    except Exception as e:  # noqa: BLE001
        return f"ERRO ao executar {nome}: {e}"
